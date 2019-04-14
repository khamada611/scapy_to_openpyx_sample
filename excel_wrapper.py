#!/usr/bin/env python
# coding: shift-jis

import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image


class ExcelWapper:
    """
    This is sample code.
    This class create new excel file and create one sheet.
    This class can access under "Z" colomn only.
    This class can't do anything else..
    """

    def __init__(self):
        """Initialize
        """
        self.book = None
        self.sheet = None
        self.title_color = "c0eec0"
        self.data_color = "f5f5f5"
        self.title_fill = PatternFill(
            patternType='solid',
            fgColor=self.title_color,
            bgColor=self.title_color
        )
        self.data_fill = PatternFill(
            patternType='solid',
            fgColor=self.data_color,
            bgColor=self.data_color
        )
        self.border = Border(top=Side(style='thin', color='000000'),
                             bottom=Side(style='thin', color='000000'),
                             left=Side(style='thin', color='000000'),
                             right=Side(style='thin', color='000000')
                             )

    def __get_xpos(self, x):
        """PRIVATE: Get excel colmn position
        """
        return chr(ord("A") - 1 + x)

    def create_book(self):
        """Create openpyxc WorkBook.

        """
        self.book = openpyxl.Workbook()

    def create_sheet(self, name):
        """Create excel sheet.

        Parameters
        ----------
        name : str
            Sheet name
        """
        self.sheet = self.book.create_sheet(name, 0)

    def write_value(self, x, y, value):
        """Write value to sheet.

        Parameters
        ----------
        x : int
            x pos (A=1, B=2,  ...)
        y : int
            y pos (1, 2, ...)
        value : 
            value(I guess it is double or int or string ..?) 
        """
        self.sheet.cell(row=y, column=x, value=value)

    def resize_sheet_width(self):
        """Resize your sheet.

        Resize width of all valid cell 

        """
        for column in self.sheet.columns:
            max_width = 0
            x = column[0].column
            for cell in column:
                now_width = len(str(cell.value))
                if now_width > max_width:
                    max_width = now_width
            if max_width == 0:
                continue
            max_width += 5
            pos = self.__get_xpos(x)
            self.sheet.column_dimensions[pos].width = max_width

    def draw_table(self, x, x_size, y, y_size):
        """Draw table by PatternFill and Border

        Parameters
        ----------
        x : int
            x pos (A=1, B=2,  ...)
        x_size : int
            num of cell
        y : int
            y pos (1, 2, ...)
        y_size : int
            num of cell
        """
        title_y = y
        for col_num in range(x, x + x_size):
            for row_num in range(y, y + y_size):
                if row_num == title_y:
                    self.sheet.cell(
                        row=row_num, column=col_num).fill = self.title_fill
                else:
                    self.sheet.cell(
                        row=row_num, column=col_num).fill = self.data_fill
                self.sheet.cell(
                    row=row_num, column=col_num).border = self.border

    def __calc_length_of_image(self, sheet_length, img_length, is_width):
        """PRIVATE: Calc length of image for excell format
        """
        if sheet_length != None and sheet_length >= img_length:
            return sheet_length
        if is_width == True:
            return img_length / 8
        else:  # height
            return img_length / 1.325

    def add_imagefile(self, x, y, img_filename, resize_x, resize_y):
        """Add image file to cell.

        Parameters
        ----------
        x : int
            x pos (A=1, B=2,  ...)
        y : int
            y pos (1, 2, ...)
        img_filename : str
            Image file's path
        resize_x : Bool
            If you need to resize X pos(A, B, C, ..) set True
        resize_y : Bool
            If you need to resize Y pos(1, 2, 3 ..) set True
        """
        x_pos = self.__get_xpos(x)
        xy_pos = x_pos + str(y)
        img = Image(img_filename)
        self.sheet.add_image(img, xy_pos)

        if resize_x == True:
            self.sheet.column_dimensions[x_pos].width = self.__calc_length_of_image(
                self.sheet.column_dimensions[x_pos].width, img.width, True)

        if resize_y == True:
            self.sheet.row_dimensions[y].height = self.__calc_length_of_image(
                self.sheet.row_dimensions[y].height, img.height, False)

    def save(self, name):
        """Save excel file.

        Parameters
        ----------
        name : str
            File name
        """
        self.book.save(name)
        return
